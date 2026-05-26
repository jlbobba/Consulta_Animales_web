import csv
import base64
import hashlib
import hmac
import io
import json
import os
import secrets
import sqlite3
import time
import traceback
from datetime import datetime
from html import escape
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlencode, urlparse


BASE_DIR = Path(__file__).resolve().parent
DB_FILE = Path(os.environ.get("DB_PATH", BASE_DIR / "animales.db"))
DATABASE_URL = os.environ.get("DATABASE_URL")
APP_USERS = os.environ.get("APP_USERS", "")
SESSION_SECRET = os.environ.get("SESSION_SECRET", "")
SESSION_COOKIE = "consulta_animales_session"
SESSION_MAX_AGE = 12 * 60 * 60


def b64_encode(data):
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def b64_decode(data):
    padding = "=" * (-len(data) % 4)
    return base64.urlsafe_b64decode((data + padding).encode("ascii"))


def auth_config_error():
    if using_postgres() and not APP_USERS:
        return "Falta configurar APP_USERS en las variables de entorno."
    if using_postgres() and not SESSION_SECRET:
        return "Falta configurar SESSION_SECRET en las variables de entorno."
    return None


def session_secret():
    if SESSION_SECRET:
        return SESSION_SECRET
    return "local-dev-secret"


def load_users():
    raw_users = APP_USERS
    if not raw_users and not using_postgres():
        raw_users = "consulta:consulta:consulta,admin:admin:editor"

    users = {}
    for item in raw_users.split(","):
        parts = item.strip().split(":", 2)
        if len(parts) != 3:
            continue
        username, password, role = [part.strip() for part in parts]
        if username and password and role in {"consulta", "editor"}:
            users[username] = {"password": password, "role": role}
    return users


def sign_payload(payload):
    data = b64_encode(json.dumps(payload, separators=(",", ":")).encode("utf-8"))
    signature = hmac.new(session_secret().encode("utf-8"), data.encode("ascii"), hashlib.sha256).hexdigest()
    return f"{data}.{signature}"


def verify_session(cookie_value):
    if not cookie_value or "." not in cookie_value:
        return None
    data, signature = cookie_value.rsplit(".", 1)
    expected = hmac.new(session_secret().encode("utf-8"), data.encode("ascii"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(signature, expected):
        return None
    try:
        payload = json.loads(b64_decode(data).decode("utf-8"))
    except Exception:
        return None
    if payload.get("exp", 0) < time.time():
        return None
    username = payload.get("username")
    role = payload.get("role")
    if not username or role not in {"consulta", "editor"}:
        return None
    return {"username": username, "role": role}


def user_can_edit(user):
    return bool(user and user.get("role") == "editor")


def using_postgres():
    return bool(DATABASE_URL)


def db_connect():
    if using_postgres():
        try:
            import psycopg
            from psycopg.rows import dict_row
        except ImportError as exc:
            raise RuntimeError(
                "Para usar Supabase/PostgreSQL instale dependencias con: "
                "python3 -m pip install -r requirements.txt"
            ) from exc
        return psycopg.connect(DATABASE_URL, row_factory=dict_row)

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def db_sql(query):
    if using_postgres():
        return query.replace("?", "%s")
    return query


def execute(conn, query, params=()):
    return conn.execute(db_sql(query), params)


def get_actions():
    with db_connect() as conn:
        rows = execute(conn, 'SELECT "Accion" FROM "Acciones" ORDER BY "Accion"').fetchall()
    return [row["Accion"] for row in rows]


def date_clauses(column, start, end):
    clauses = []
    params = []
    start = (start or "").strip()
    end = (end or "").strip()

    if start:
        clauses.append(f"{column} >= ?")
        params.append(start)

    if end:
        if len(end) in (4, 7):
            clauses.append(f"{column} <= (? || 'Z')")
            params.append(end)
        else:
            clauses.append(f"{column} <= ?")
            params.append(end)

    return clauses, params


def movement_exists_sql(table, alias, date_column, start, end):
    clauses, params = date_clauses(date_column, start, end)
    where = f'{alias}."IDE" = d."IDE"'
    for clause in clauses:
        where += f" AND {clause}"
    return f'EXISTS (SELECT 1 FROM "{table}" {alias} WHERE {where})', params


def clean_filter_params(params):
    allowed = {
        "ide",
        "fecha_desde",
        "fecha_hasta",
        "fecha_mov_desde",
        "fecha_mov_hasta",
        "accion",
        "sexo",
        "only_with_weights",
        "sort_by",
        "sort_dir",
    }
    return {key: value for key, value in params.items() if key in allowed and value}


def view_url(view, filters):
    params = clean_filter_params(filters)
    params.pop("sort_by", None)
    params.pop("sort_dir", None)
    params["view"] = view
    return "/?" + urlencode(params)


def path_with_message(path, message, message_type):
    parsed = urlparse(path or "/")
    params = {key: values[0] for key, values in parse_qs(parsed.query).items()}
    params["message"] = message
    params["type"] = message_type
    return (parsed.path or "/") + "?" + urlencode(params)


def parse_number(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def parse_iso_date(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        return None


def latest_weight_by_ide(rows):
    latest = {}
    for row in rows:
        ide = row["IDE"]
        weight = parse_number(row["Peso"])
        if ide is None or weight is None:
            continue
        date_value = parse_iso_date(row["Fecha"]) or datetime.min
        time_value = str(row["Hora"] or "")
        key = str(ide)
        current = latest.get(key)
        if current is None or (date_value, time_value) > current["sort_key"]:
            latest[key] = {"sort_key": (date_value, time_value), "weight": weight}
    return [item["weight"] for item in latest.values()]


def applied_filters_text(filters):
    labels = {
        "ide": "IDE contiene",
        "fecha_desde": "FNac. desde",
        "fecha_hasta": "FNac. hasta",
        "fecha_mov_desde": "Mov. desde",
        "fecha_mov_hasta": "Mov. hasta",
        "sexo": "Sexo",
        "accion": "Acción",
    }
    parts = []
    for key in ["ide", "fecha_desde", "fecha_hasta", "fecha_mov_desde", "fecha_mov_hasta", "sexo", "accion"]:
        value = (filters.get(key) or "").strip()
        if not value or value == "Todos":
            continue
        separator = " " if key == "ide" else " = "
        parts.append(f"{labels[key]}{separator}{value}")

    if filters.get("only_with_weights") == "1":
        parts.append("Solo con pesajes")

    if not parts:
        return "Filtros aplicados: Ninguno"
    return "Filtros aplicados: " + " | ".join(parts)


def add_pesos_calculations(cols, rows):
    if not rows:
        return cols, rows

    calculated_rows = []
    previous_by_ide = {}
    for row in rows:
        item = {col: row[col] for col in cols}
        ide = item.get("IDE")
        peso = parse_number(item.get("Peso"))
        fecha = parse_iso_date(item.get("Fecha"))
        gdm_calc = ""

        if ide is not None and peso is not None and fecha is not None:
            previous = previous_by_ide.get(str(ide))
            if previous:
                previous_fecha, previous_peso = previous
                days = (fecha - previous_fecha).days
                if days > 0:
                    gdm_calc = f"{((peso - previous_peso) / days):.3f}"
            previous_by_ide[str(ide)] = (fecha, peso)

        item["GMD_Calc"] = gdm_calc
        calculated_rows.append(item)

    if "GMD_Calc" not in cols:
        if "GDM" in cols:
            insert_at = cols.index("GDM") + 1
            cols = cols[:insert_at] + ["GMD_Calc"] + cols[insert_at:]
        else:
            cols = cols + ["GMD_Calc"]
    return cols, calculated_rows


def sort_value(value):
    if value is None or str(value).strip() == "":
        return None

    number = parse_number(value)
    if number is not None:
        return (0, number)

    date_value = parse_iso_date(value)
    if date_value is not None:
        return (1, date_value)

    return (2, str(value).lower())


def sort_rows(cols, rows, filters):
    sort_by = filters.get("sort_by")
    sort_dir = filters.get("sort_dir", "asc")
    if sort_by not in cols:
        return rows

    with_values = []
    without_values = []
    for row in rows:
        value = sort_value(row[sort_by])
        if value is None:
            without_values.append(row)
        else:
            with_values.append((value, row))

    with_values.sort(key=lambda item: item[0], reverse=sort_dir == "desc")
    return [row for _, row in with_values] + without_values


def run_query(view, filters):
    params = []

    if view == "pesos":
        join_type = "JOIN" if filters.get("only_with_weights") == "1" else "LEFT JOIN"
        query = f"""
            SELECT
                d."IDE",
                d."IDV",
                d."SEXO",
                d."RAZA",
                d."CRUZA",
                d."FECHANAC",
                p."Peso",
                p."Fecha",
                p."Hora",
                p."GDM",
                p."Nota"
            FROM "DatoAnimal" d
            {join_type} "Pesos" p ON d."IDE" = p."IDE"
            WHERE 1=1
        """
        order_by = ' ORDER BY d."IDE", p."Fecha", p."Hora"'
    elif view == "acciones":
        query = """
            SELECT
                d."IDE",
                d."IDV",
                d."SEXO",
                d."RAZA",
                d."CRUZA",
                d."FECHANAC",
                l."Fecha",
                l."Accion"
            FROM "DatoAnimal" d
            LEFT JOIN "Lecturas" l ON d."IDE" = l."IDE"
            WHERE 1=1
        """
        order_by = ' ORDER BY d."IDE", l."Fecha"'
    else:
        view = "animales"
        query = """
            SELECT
                d."IDE",
                d."IDV",
                d."SEXO",
                d."RAZA",
                d."CRUZA",
                d."FECHANAC",
                (
                    SELECT p."Peso"
                    FROM "Pesos" p
                    WHERE p."IDE" = d."IDE"
                    ORDER BY p."Fecha" DESC, p."Hora" DESC
                    LIMIT 1
                ) AS "UltimoPeso",
                (
                    SELECT p."Fecha"
                    FROM "Pesos" p
                    WHERE p."IDE" = d."IDE"
                    ORDER BY p."Fecha" DESC, p."Hora" DESC
                    LIMIT 1
                ) AS "FechaUltimoPeso",
                (
                    SELECT l."Accion"
                    FROM "Lecturas" l
                    WHERE l."IDE" = d."IDE"
                    ORDER BY l."Fecha" DESC, l."ID" DESC
                    LIMIT 1
                ) AS "UltimaAccion",
                (
                    SELECT l."Fecha"
                    FROM "Lecturas" l
                    WHERE l."IDE" = d."IDE"
                    ORDER BY l."Fecha" DESC, l."ID" DESC
                    LIMIT 1
                ) AS "FechaUltimaAccion"
            FROM "DatoAnimal" d
            WHERE 1=1
        """
        order_by = ' ORDER BY d."IDE"'

    ide = (filters.get("ide") or "").strip()
    if ide:
        query += ' AND CAST(d."IDE" AS TEXT) LIKE ?'
        params.append(f"%{ide}%")

    sexo = (filters.get("sexo") or "").strip()
    if sexo and sexo != "Todos":
        query += ' AND d."SEXO" = ?'
        params.append(sexo)

    clauses, clause_params = date_clauses(
        'd."FECHANAC"',
        filters.get("fecha_desde"),
        filters.get("fecha_hasta"),
    )
    for clause in clauses:
        query += f" AND {clause}"
    params.extend(clause_params)

    if view == "pesos":
        mov_clauses, mov_params = date_clauses(
            'p."Fecha"',
            filters.get("fecha_mov_desde"),
            filters.get("fecha_mov_hasta"),
        )
        for clause in mov_clauses:
            query += f" AND {clause}"
        params.extend(mov_params)
    elif view == "acciones":
        mov_clauses, mov_params = date_clauses(
            'l."Fecha"',
            filters.get("fecha_mov_desde"),
            filters.get("fecha_mov_hasta"),
        )
        for clause in mov_clauses:
            query += f" AND {clause}"
        params.extend(mov_params)
    elif filters.get("fecha_mov_desde") or filters.get("fecha_mov_hasta"):
        peso_exists, peso_params = movement_exists_sql(
            "Pesos",
            "pm",
            'pm."Fecha"',
            filters.get("fecha_mov_desde"),
            filters.get("fecha_mov_hasta"),
        )
        accion_exists, accion_params = movement_exists_sql(
            "Lecturas",
            "lm",
            'lm."Fecha"',
            filters.get("fecha_mov_desde"),
            filters.get("fecha_mov_hasta"),
        )
        query += f" AND ({peso_exists} OR {accion_exists})"
        params.extend(peso_params)
        params.extend(accion_params)

    action = filters.get("accion")
    if view == "acciones" and action and action != "Todos":
        query += ' AND l."Accion" = ?'
        params.append(action)
    elif action and action != "Todos":
        query += ' AND EXISTS (SELECT 1 FROM "Lecturas" lf WHERE lf."IDE" = d."IDE" AND lf."Accion" = ?)'
        params.append(action)

    if view != "pesos" and filters.get("only_with_weights") == "1":
        query += ' AND EXISTS (SELECT 1 FROM "Pesos" pf WHERE pf."IDE" = d."IDE")'

    with db_connect() as conn:
        rows = execute(conn, query + order_by, params).fetchall()

    cols = list(rows[0].keys()) if rows else []
    if view == "pesos":
        cols, rows = add_pesos_calculations(cols, rows)
    rows = sort_rows(cols, rows, filters)
    return view, list(cols), rows


def get_idv_for_ide(ide):
    with db_connect() as conn:
        row = execute(conn, 'SELECT "IDV" FROM "DatoAnimal" WHERE "IDE" = ?', (ide,)).fetchone()
    return row["IDV"] if row else None


def add_action(ide, action, date_text):
    if not ide or not action or not date_text:
        return False, "Debe completar IDE, acción y fecha."

    idv = get_idv_for_ide(ide)
    if idv is None:
        return False, f"No se encontró IDV para IDE {ide}."

    with db_connect() as conn:
        execute(
            conn,
            'INSERT INTO "Lecturas" ("IDE", "IDV", "Fecha", "Accion") VALUES (?, ?, ?, ?)',
            (ide, idv, date_text, action),
        )
        conn.commit()
    return True, "Acción registrada correctamente."


def change_ide(current_ide, new_ide, date_text):
    if not current_ide or not new_ide or not date_text:
        return False, "Debe completar IDE actual, IDE nuevo y fecha."

    try:
        datetime.strptime(date_text, "%Y-%m-%d")
    except ValueError:
        return False, "La fecha debe tener formato YYYY-MM-DD."

    with db_connect() as conn:
        try:
            execute(conn, "BEGIN;")
            exists = execute(conn, 'SELECT COUNT(*) AS total FROM "DatoAnimal" WHERE "IDE" = ?', (current_ide,)).fetchone()
            if exists["total"] == 0:
                raise ValueError("El IDE actual no existe.")

            duplicated = execute(conn, 'SELECT COUNT(*) AS total FROM "DatoAnimal" WHERE "IDE" = ?', (new_ide,)).fetchone()
            if duplicated["total"] > 0:
                raise ValueError("El IDE nuevo ya existe.")

            execute(conn, 'UPDATE "DatoAnimal" SET "IDE" = ? WHERE "IDE" = ?', (new_ide, current_ide))
            execute(conn, 'UPDATE "Pesos" SET "IDE" = ? WHERE "IDE" = ?', (new_ide, current_ide))
            execute(conn, 'UPDATE "Lecturas" SET "IDE" = ? WHERE "IDE" = ?', (new_ide, current_ide))
            execute(
                conn,
                'INSERT INTO "Caravanas" ("IDEi", "IDEf", "Fecha") VALUES (?, ?, ?)',
                (current_ide, new_ide, date_text),
            )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
    return True, f"IDE {current_ide} cambiado a {new_ide}."


def export_csv(cols, rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(cols)
    for row in rows:
        writer.writerow(["" if row[col] is None else row[col] for col in cols])
    return output.getvalue().encode("utf-8-sig")


def format_decimal_value(value, decimal_sep):
    number = parse_number(value)
    if number is None:
        return "" if value is None else str(value)
    text = f"{number:.3f}".rstrip("0").rstrip(".")
    if decimal_sep == ",":
        text = text.replace(".", ",")
    return text


def export_excel(cols, rows, decimal_sep="."):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("Falta instalar openpyxl para exportar Excel.") from exc

    numeric_decimal_cols = {"Peso", "GDM", "GMD_Calc"}
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row_index, row in enumerate(rows, start=2):
        for col_index, col in enumerate(cols, start=1):
            value = row[col]
            cell = ws.cell(row=row_index, column=col_index)
            if col in numeric_decimal_cols:
                cell.value = format_decimal_value(value, decimal_sep)
                cell.number_format = "@"
            else:
                cell.value = "" if value is None else value

    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 28)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def html_page(title, body, message=None, message_type="ok", active_view="animales", user=None, filters=None):
    flash = ""
    if message:
        flash = f'<div class="flash {escape(message_type)}">{escape(message)}</div>'
    filters = filters or {}
    active = {
        "animales": "active" if active_view == "animales" else "",
        "pesos": "active" if active_view == "pesos" else "",
        "acciones": "active" if active_view == "acciones" else "",
    }
    user_bar = ""
    if user:
        role_label = "Editor" if user.get("role") == "editor" else "Consulta"
        user_bar = f"""
    <div class="userbar">
      <span>{escape(user["username"])} · {escape(role_label)}</span>
      <form method="post" action="/logout"><button type="submit" class="secondary-button">Salir</button></form>
    </div>
"""
    return f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(title)}</title>
  <link rel="stylesheet" href="/static/styles.css">
</head>
<body>
  <header>
    <h1>Consultas de Animales</h1>
    <nav>
      <a class="{active["animales"]}" href="{escape(view_url("animales", filters))}">Animales</a>
      <a class="{active["pesos"]}" href="{escape(view_url("pesos", filters))}">Pesos</a>
      <a class="{active["acciones"]}" href="{escape(view_url("acciones", filters))}">Acciones</a>
    </nav>
    {user_bar}
  </header>
  <main>
    {flash}
    {body}
  </main>
</body>
</html>"""


def login_page(message=None):
    config_error = auth_config_error()
    if config_error:
        body = f"""
<section class="login-panel">
  <h2>Configuración pendiente</h2>
  <p>{escape(config_error)}</p>
</section>
"""
        return html_page("Login", body)

    flash = f'<div class="flash error">{escape(message)}</div>' if message else ""
    body = f"""
{flash}
<section class="login-panel">
  <h2>Ingresar</h2>
  <form method="post" action="/login">
    <label>Usuario <input name="username" autocomplete="username" required></label>
    <label>Contraseña <input name="password" type="password" autocomplete="current-password" required></label>
    <button type="submit">Entrar</button>
  </form>
</section>
"""
    return html_page("Login", body)


def filters_form(view, filters, actions):
    selected = lambda value: "selected" if filters.get("accion", "Todos") == value else ""
    selected_sexo = lambda value: "selected" if filters.get("sexo", "Todos") == value else ""
    checked = "checked" if filters.get("only_with_weights") == "1" else ""
    action_options = "\n".join(
        f'<option value="{escape(action)}" {selected(action)}>{escape(action)}</option>'
        for action in ["Todos"] + actions
    )
    return f"""
<form class="filters" method="get" action="/">
  <input type="hidden" name="view" value="{escape(view)}">
  <label>IDE
    <input name="ide" value="{escape(filters.get("ide", ""))}" placeholder="Busca parcial">
  </label>
  <label>FNac. desde
    <input name="fecha_desde" value="{escape(filters.get("fecha_desde", ""))}" placeholder="YYYY, YYYY-MM o YYYY-MM-DD">
  </label>
  <label>FNac. hasta
    <input name="fecha_hasta" value="{escape(filters.get("fecha_hasta", ""))}" placeholder="YYYY, YYYY-MM o YYYY-MM-DD">
  </label>
  <label>Mov. desde
    <input name="fecha_mov_desde" value="{escape(filters.get("fecha_mov_desde", ""))}" placeholder="YYYY, YYYY-MM o YYYY-MM-DD">
  </label>
  <label>Mov. hasta
    <input name="fecha_mov_hasta" value="{escape(filters.get("fecha_mov_hasta", ""))}" placeholder="YYYY, YYYY-MM o YYYY-MM-DD">
  </label>
  <label>Sexo
    <select name="sexo">
      <option value="Todos" {selected_sexo("Todos")}>Todos</option>
      <option value="Hembra" {selected_sexo("Hembra")}>Hembra</option>
      <option value="Macho" {selected_sexo("Macho")}>Macho</option>
    </select>
  </label>
  <label class="acciones">Acción
    <select name="accion">{action_options}</select>
  </label>
  <label class="checkbox">
    <input type="checkbox" name="only_with_weights" value="1" {checked}>
    Solo con pesajes
  </label>
  <button type="submit">Consultar</button>
  <a class="button secondary" href="/?view={escape(view)}">Limpiar</a>
</form>
"""


def table_html(cols, rows, filters, view):
    if not cols:
        return f"""
<section class="results-summary">
  <div class="filters-applied">{escape(applied_filters_text(filters))}</div>
</section>
<section class="empty">No hay resultados para mostrar.</section>
"""

    export_params = {k: v for k, v in filters.items() if v and k not in {"message", "type"}}
    export_params["view"] = view
    csv_params = dict(export_params)
    csv_params["format"] = "csv"
    xlsx_params = dict(export_params)
    xlsx_params["format"] = "xlsx"
    csv_link = f"/export?{escape(urlencode(csv_params))}"
    summary_parts = [f"Total filas: {len(rows)}"]
    if view == "pesos" and "Peso" in cols:
        weights = latest_weight_by_ide(rows)
        summary_parts.append(f"Total animales con peso: {len(weights)}")
        if weights:
            summary_parts.append(f"Suma ultimos pesos: {sum(weights):.1f}")
            summary_parts.append(f"Peso promedio: {(sum(weights) / len(weights)):.1f}")
    elif view == "animales" and "UltimoPeso" in cols:
        weights = [parse_number(row["UltimoPeso"]) for row in rows]
        weights = [weight for weight in weights if weight is not None]
        summary_parts.append(f"Total con peso: {len(weights)}")
        if weights:
            summary_parts.append(f"Suma pesos: {sum(weights):.1f}")
            summary_parts.append(f"Peso promedio: {(sum(weights) / len(weights)):.1f}")
    summary = " | ".join(summary_parts)
    filters_summary = applied_filters_text(filters)
    active_sort = filters.get("sort_by")
    active_dir = filters.get("sort_dir", "asc")
    head_cells = []
    for col in cols:
        next_dir = "desc" if active_sort == col and active_dir != "desc" else "asc"
        sort_params = dict(export_params)
        sort_params["sort_by"] = col
        sort_params["sort_dir"] = next_dir
        marker = " ▲" if active_sort == col and active_dir != "desc" else " ▼" if active_sort == col else ""
        head_cells.append(
            f'<th><a class="sort-link" href="/?{escape(urlencode(sort_params))}">{escape(col)}{marker}</a></th>'
        )
    head = "".join(head_cells)
    body_rows = []
    for row in rows:
        cells = "".join(f"<td>{escape('' if row[col] is None else str(row[col]))}</td>" for col in cols)
        body_rows.append(f"<tr>{cells}</tr>")
    return f"""
<section class="results-summary">
  <div class="filters-applied">{escape(filters_summary)}</div>
  <div class="results-head">
    <strong>{escape(summary)}</strong>
  <div class="export-actions">
    <a class="button secondary" href="{csv_link}">Exportar CSV</a>
    <form class="export-excel-form" method="get" action="/export">
      <input type="hidden" name="format" value="xlsx">
      {"".join(f'<input type="hidden" name="{escape(k)}" value="{escape(v)}">' for k, v in xlsx_params.items() if k != "format")}
      <button type="submit">Exportar Excel</button>
      <label>Decimal
        <select name="decimal_sep">
          <option value=".">Punto</option>
          <option value=",">Coma</option>
        </select>
      </label>
    </form>
  </div>
  </div>
</section>
<div class="table-wrap">
  <table>
    <thead><tr>{head}</tr></thead>
    <tbody>{''.join(body_rows)}</tbody>
  </table>
</div>
"""


def action_form(actions, user):
    if not user_can_edit(user):
        return """
<section class="panel muted">
  <h2>Agregar acción</h2>
  <p>Disponible solo para usuarios editores.</p>
</section>
"""
    options = "".join(f'<option value="{escape(action)}">{escape(action)}</option>' for action in actions)
    today = datetime.now().strftime("%d/%m/%Y")
    return f"""
<section class="panel">
  <h2>Agregar acción</h2>
  <form method="post" action="/add-action">
    <label>IDE <input name="ide" required></label>
    <label>Acción <select name="accion" required>{options}</select></label>
    <label>Fecha <input name="fecha" value="{today}" required></label>
    <button type="submit">Guardar acción</button>
  </form>
</section>
"""


def change_ide_form(user):
    if not user_can_edit(user):
        return """
<section class="panel muted">
  <h2>Cambiar IDE</h2>
  <p>Disponible solo para usuarios editores.</p>
</section>
"""
    today = datetime.now().strftime("%Y-%m-%d")
    return f"""
<section class="panel">
  <h2>Cambiar IDE</h2>
  <form method="post" action="/change-ide">
    <label>IDE actual <input name="ide_actual" required></label>
    <label>IDE nuevo <input name="ide_nuevo" required></label>
    <label>Fecha <input name="fecha" value="{today}" required></label>
    <button type="submit">Cambiar IDE</button>
  </form>
</section>
"""


class AppHandler(BaseHTTPRequestHandler):
    def do_HEAD(self):
        if urlparse(self.path).path == "/":
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            return
        self.send_response(HTTPStatus.NOT_FOUND)
        self.end_headers()

    def send_html(self, html, status=HTTPStatus.OK):
        data = html.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_html_with_cookie(self, html, cookie, status=HTTPStatus.OK):
        data = html.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Set-Cookie", cookie)
        self.end_headers()
        self.wfile.write(data)

    def redirect(self, path):
        self.send_response(HTTPStatus.SEE_OTHER)
        self.send_header("Location", path)
        self.end_headers()

    def redirect_with_cookie(self, path, cookie):
        self.send_response(HTTPStatus.SEE_OTHER)
        self.send_header("Location", path)
        self.send_header("Set-Cookie", cookie)
        self.end_headers()

    def redirect_back(self, message, message_type="ok"):
        referer = self.headers.get("Referer", "/")
        parsed = urlparse(referer)
        path = parsed.path or "/"
        if parsed.query:
            path += "?" + parsed.query
        self.redirect(path_with_message(path, message, message_type))

    def current_user(self):
        cookie_header = self.headers.get("Cookie", "")
        cookies = {}
        for item in cookie_header.split(";"):
            if "=" in item:
                key, value = item.strip().split("=", 1)
                cookies[key] = value
        return verify_session(cookies.get(SESSION_COOKIE))

    def read_form(self):
        length = int(self.headers.get("Content-Length", "0"))
        data = self.rfile.read(length).decode("utf-8")
        return {key: values[0] for key, values in parse_qs(data).items()}

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            params = {key: values[0] for key, values in parse_qs(parsed.query).items()}

            if parsed.path == "/static/styles.css":
                css = (BASE_DIR / "static" / "styles.css").read_bytes()
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "text/css; charset=utf-8")
                self.send_header("Content-Length", str(len(css)))
                self.end_headers()
                self.wfile.write(css)
                return

            if parsed.path == "/login":
                self.send_html(login_page(params.get("message")))
                return

            user = self.current_user()
            if not user:
                self.redirect("/login")
                return

            if parsed.path == "/export":
                view, cols, rows = run_query(params.get("view", "animales"), params)
                export_format = params.get("format", "csv")
                if export_format == "xlsx":
                    data = export_excel(cols, rows, params.get("decimal_sep", "."))
                    self.send_response(HTTPStatus.OK)
                    self.send_header(
                        "Content-Type",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    self.send_header("Content-Disposition", f'attachment; filename="{view}.xlsx"')
                    self.send_header("Content-Length", str(len(data)))
                    self.end_headers()
                    self.wfile.write(data)
                else:
                    data = export_csv(cols, rows)
                    self.send_response(HTTPStatus.OK)
                    self.send_header("Content-Type", "text/csv; charset=utf-8")
                    self.send_header("Content-Disposition", f'attachment; filename="{view}.csv"')
                    self.send_header("Content-Length", str(len(data)))
                    self.end_headers()
                    self.wfile.write(data)
                return

            if parsed.path != "/":
                self.send_html(html_page("No encontrado", "<p>No encontrado.</p>"), HTTPStatus.NOT_FOUND)
                return

            actions = get_actions()
            view, cols, rows = run_query(params.get("view", "animales"), params)
            message = params.get("message")
            message_type = params.get("type", "ok")
            current_filters = clean_filter_params(params)
            body = (
                filters_form(view, params, actions)
                + table_html(cols, rows, params, view)
                + '<section class="forms-grid">'
                + action_form(actions, user)
                + change_ide_form(user)
                + "</section>"
            )
            self.send_html(html_page("Consultas de Animales", body, message, message_type, view, user, current_filters))
        except Exception as exc:
            traceback.print_exc()
            body = f"""
<section class="empty">
  <h2>Error al consultar la base</h2>
  <p>{escape(str(exc))}</p>
</section>
"""
            self.send_html(html_page("Error", body), HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_POST(self):
        try:
            form = self.read_form()
            if self.path == "/login":
                users = load_users()
                username = form.get("username", "").strip()
                password = form.get("password", "")
                user_record = users.get(username)
                if not user_record or not secrets.compare_digest(password, user_record["password"]):
                    self.redirect("/login?" + urlencode({"message": "Usuario o contraseña incorrectos."}))
                    return
                payload = {
                    "username": username,
                    "role": user_record["role"],
                    "exp": int(time.time() + SESSION_MAX_AGE),
                }
                cookie = (
                    f"{SESSION_COOKIE}={sign_payload(payload)}; Path=/; HttpOnly; "
                    f"SameSite=Lax; Max-Age={SESSION_MAX_AGE}"
                )
                self.redirect_with_cookie("/", cookie)
                return

            if self.path == "/logout":
                cookie = f"{SESSION_COOKIE}=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0"
                self.redirect_with_cookie("/login", cookie)
                return

            user = self.current_user()
            if not user:
                self.redirect("/login")
                return

            if self.path == "/add-action":
                if not user_can_edit(user):
                    self.redirect_back("No tiene permiso para agregar acciones.", "error")
                    return
                ok, message = add_action(form.get("ide"), form.get("accion"), form.get("fecha"))
            elif self.path == "/change-ide":
                if not user_can_edit(user):
                    self.redirect_back("No tiene permiso para cambiar IDE.", "error")
                    return
                ok, message = change_ide(form.get("ide_actual"), form.get("ide_nuevo"), form.get("fecha"))
            else:
                self.send_html(html_page("No encontrado", "<p>No encontrado.</p>"), HTTPStatus.NOT_FOUND)
                return
            self.redirect_back(message, "ok" if ok else "error")
        except Exception as exc:
            self.redirect_back(str(exc), "error")


def main():
    port = int(os.environ.get("PORT", "8000"))
    host = os.environ.get("HOST", "127.0.0.1")
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"App web lista en http://{host}:{port}")
    print(f"Base de datos: {'Supabase/PostgreSQL' if using_postgres() else DB_FILE}")
    server.serve_forever()


if __name__ == "__main__":
    main()
